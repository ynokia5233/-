# /usr/bin/env python
# -*- coding:utf-8 -*-


__author__ = 'zWX537642'

import openpyxl.utils
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side
import json
#import sys
#reload(sys)
#sys.setdefaultencoding("utf-8")

class Openpyxl_Excel(object):
    def __init__(self):
        pass

    def openpyxl_OpenExcel_DataFalse(self, url, read=False):

        """
        打开一个已有的Excel表格
        :param url: 打开的Excel路径
        :param read: 只读属性，默认为True
        :return:可读取的是单元格的公式

        """
        self.wb = load_workbook(url, read_only=read)


    def openpyxl_OpenExcel(self, url, read=False, data=True):

        """
        打开一个已有的Excel表格
        :param url: 打开的Excel路径
        :param read: 只读属性，默认为True
        :return:读取的是单元格的具体数据
        """
        self.wb = load_workbook(url, read_only=read, data_only=data)

    def openpyxl_Get_ExcelReadOnly(self):
        """
        :return:返回Excel的read_only属性
        """
        return self.wb.read_only()

    def openpyxl_Get_ExcelDataOnly(self):
        """
        :return:返回Excel的data_only属性
        """
        return self.wb.data_only()

    def openpyxl_Get_ExcelKeepLinks(self):
        """
        :return:返回Excel的keep_links属性
        """
        return self.wb.keep_links()

    def openpyxl_CloseExcel(self):
        """
        关闭打开的Excel表格
        :return:
        """
        self.wb.close()

    def openpyxl_SaveExcel(self, filename):
        """
        保存Excel表格
        :param filename: 保存路径以及Excel名称
        :return:
        """
        self.wb.save(filename)

    def openpyxl_Get_SheetNames(self, index=None):
        """
        获取Excel中的全部sheet页名称，并以list返回
        :param index:为空以list返回全部sheet名称，index不为空，如为1，则返回第二个sheet页的名称
        :return:以list返回全部sheet页名称
        """
        #return self.wb.sheetnames  #返回所有sheet名称，两种方法一样
        if index == None:
            return self.wb.get_sheet_names()
        else:
            return self.wb.get_sheet_names()[index]

    def openpyxl_Get_SheetActives(self):
        """
        获取正在使用的sheet页面对象
        :return:返回sheet对象
        """
        self.sheet = self.wb.get_active_sheet()
        return self.sheet

    def openpyxl_Get_SheetActive(self):
        """
        获取正在使用的sheet对象
        :return:返回sheet对象
        """
        self.sheet = self.wb.active()
        return self.sheet

    def openpyxl_Get_Sheet_ByName(self, sheetName):
        """
        通过sheet的名称获取sheet对象
        :param sheetName: sheetName的名称
        :return:返回sheet对象
        """
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        return self.sheet

    def openpyxl_Get_Sheet_ByIndex(self, index=None):
        """
        根据sheet的索引获取sheet对象
        :param index: 索引,如果index==None,以list返回所有sheet对象
        :return:返回sheet对象
        """
        if index ==None:
            self.sheet = self.wb.worksheets
            return self.sheet
        else:
            self.sheet = self.wb.worksheets[index]
            return self.sheet

    def openpyxl_Get_SheetTitle(self, sheet):
        """
        获取sheet的索引值，参数为sheet对象
        :param sheet: sheet的对象
        :return:返回sheet页的名称
        """
        return sheet.title



    def openpyxl_Get_SheetIndex(self, sheet):
        """
        获取sheet的索引值，参数为sheet对象
        :param sheet: sheet的对象
        :return:返回sheet索引
        """
        #retrun self.wb.index(sheet) #两种方法都可以返回sheet的索引值，参数都为sheet对象
        return self.wb.get_index(sheet)

    def openpyxl_SheetCreate(self, title=None, index=None):
        """
        创建sheet
        :param title: 新创建的sheet名称，默认名字sheetnumber
        :param index: 从指定位置插入新的sheet，默认为最后
        :return:返回新创建的sheet对象
        """
        self.sheet = self.wb.create_sheet(title=title, index=index)
        return self.sheet

    def openpyxl_SheetCopy(self, sheet):
        """
        复制sheet，复制的sheet默认在最后
        :param sheet: 参数为sheet的对象
        :return:返回新创建的sheet对象
        """
        self.sheet = self.wb.copy_worksheet(sheet)
        return self.sheet

    def openpyxl_SheetRemove(self, sheet):
        """
        移除sheet
        :param sheet: sheet对象
        :return:
        """
        #self.wb.remove(sheet) #两种方法相同
        self.wb.remove_sheet(sheet)

    def openpyxl_SheetRemove_ByName(self, name):
        """
        移除sheet
        :param name: sheet名称
        :return:
        """
        del self.wb[name]

    def openpyxl_Get_MaxRow(self):
        """
        获取最大行
        :return:返回Excel中最大行数
        """
        return self.sheet.max_row

    def openpyxl_Get_MaxColumn(self):
        """
        获取最大列
        :return:返回Excel中最大列数
        """
        return self.sheet.max_column

    def openpyxl_Get_MinRow(self):
        """
        获取最小行
        :return:返回Excel中最小行数
        """
        return self.sheet.min_row

    def openpyxl_Get_MinColumn(self):
        """
        获取最小列
        :return:返回Excel中最小列数
        """
        return self.sheet.min_column

    def openpyxl_Get_SheetDimensions(self):
        """
        返回sheet页的有效范围，如A1:C3
        :return:
        """
        #return self.sheet.calculate_dimension() #此两种方法是一样的
        return self.sheet.dimensions

    def openpyxl_Get_CellCoordinate(self, Cell):
        """
        根据坐标返回cell行数以及列值
        :param Cell: 坐标，如A4
        :return:返回行数，返回列数
        """
        return self.sheet[Cell].row, self.sheet[Cell].column

    def openpyxl_Get_CellColumenString(self, columen_number):
        """
        根据列数返回获得字母
        :param columen_number: 列数
        :return:字母
        """
        return openpyxl.utils.get_column_letter(columen_number)

    def openpyxl_Get_CellColumenNumber(self, str_col):
        """
        根据字母获取列数
        :param str_col: 字母
        :return:列数
        """
        return openpyxl.utils.column_index_from_string(str_col)

    def openpyxl_Get_CellText_ByCoordinate(self, star_cellCoordinate=None, end_cellCoordinate=None, DataFormat='list'):
        """
        获取单元格中的内容
        1、只输入起始位置的坐标，则返回目标cell中的内容
        2、输入起始位置坐标,如A1，结束位置坐标，如B3，则返回某一区域的内容
        （其中DataFormat输入list或者不输入，以每一行内容作为一个list元素返回，如[[u'qwer', u'data1'], [2L, 40L], [3L, 40L], [4L, 50L]]）
        （DataFormat为dic，返回的是一个字典，如{'B4': 50L, 'A1': u'qwer', 'A3': 3L, 'A2': 2L, 'B1': u'data1', 'B2': 40L, 'B3': 40L, 'A4': 4L}）
        :param star_cell: 起始位置，如"A1"
        :param end_cell: 结束位置，如"B4"
        :param DataFormat: 返回的数据格式
        :return:
        """
        list1 = []
        dic = {}
        if star_cellCoordinate != None and end_cellCoordinate == None:
            return self.sheet[star_cellCoordinate].value
        elif star_cellCoordinate != None and end_cellCoordinate != None and DataFormat == 'list':
            for x in self.sheet[star_cellCoordinate + ":" + end_cellCoordinate]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
            return list1
        elif star_cellCoordinate != None and end_cellCoordinate != None and DataFormat == 'dic':
            for x in self.sheet[star_cellCoordinate + ":" + end_cellCoordinate]:
                for y in x:
                        dic[y.coordinate] = y.value
            #dic = json.dumps(dic)
            #dic = dic.decode("unicode_escape")
            return dic

    def openpyxl_Get_CellText_Number(self, row, column):
        """
        获取某一单元格内容
        :param row: 行数，如1
        :param column: 列数，如3
        :return:返回单元格中的内容
        """
        return self.sheet.cell(row=row, column=column).value

    def openpyxl_Get_CellText_IterRows(self,range_string=None, min_row=None, max_row=None, min_column=None, max_column=None, DataFromat='list'):
        """
        以行获取某一区域的内容
        :param range_string:区域，如"A2:B4"
        :param min_row: 起始行
        :param max_row: 结束行
        :param min_column: 起始列
        :param max_column: 结束列
        :param DataFromat: 默认返回list，DataFormat为"dic"，返回字典
        :return:
        """
        list1 = []
        dic = {}
        if DataFromat == 'list':
            if range_string != None:
                for x in self.sheet.iter_rows(range_string=range_string):
                    list2 = []
                    for y in x:
                        list2.append(y.value)
                    list1.append(list2)
            elif range_string ==None and (min_row != None and max_row != None and min_column != None and max_column != None):
                for x in self.sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
                    list2 = []
                    for y in x:
                        list2.append(y.value)
                    list1.append(list2)
            return list1
        elif DataFromat == 'dic':
            if range_string != None:
                for x in self.sheet.iter_rows(range_string=range_string):
                    for y in x:
                        dic[y.coordinate] = y.value
            elif range_string ==None and (min_row != None and max_row != None and min_column != None and max_column != None):
                for x in self.sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
                    for y in x:
                        dic[y.coordinate] = y.value
            return dic

    def openpyxl_Get_CellText_IterColumns(self, min_column, max_column, min_row, max_row, DataFormat='list'):
        """
        以列返回某一区域中的内容
        :param min_column: 起始列
        :param max_column: 结束列
        :param min_row: 起始行
        :param max_row: 结束行
        :param DataFormat: 默认返回list，DataFormat为"dic"，返回字典
        :return:
        """
        list1 = []
        dic = {}
        if DataFormat == 'list':
            for x in self.sheet.iter_cols(min_col=min_column, max_col=max_column, min_row=min_row, max_row=max_row):
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
            return list1
        elif DataFormat == 'dic':
            for x in self.sheet.iter_cols(min_col=min_column, max_col=max_column, min_row=min_row, max_row=max_row):
                for y in x:
                    dic[y.coordinate] = y.value
            return dic

    def openpyxl_Get_CellText_Squared(self, min_col, min_row, max_col, max_row):
        """
        返回某一区域内容的
        :param min_col: 起始列
        :param min_row: 起始行
        :param max_col: 结束列
        :param max_row: 结束行
        :return:
        """
        list1 = []
        for x in self.sheet.get_squared_range(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            list2 = []
            for y in x:
                list2.append(y.value)
            list1.append(list2)
        return list1

    def openpyxl_Get_RowsText(self, star_row=None, end_row=None):
        """
        返回每行的内容，可以以切片进行处理
        1、star_row=None和end_row=None，返回sheet页全部内容
        2、star_row有值，end_row=None,返回的内容为rows[star_row:]
        3、star_row=None，end_row有值,返回的内容为rows[:end_row]
        4、star_row有值，end_row有值,返回的内容为rows[star_row:end_row]
        :param star_row: 起始行
        :param end_row: 结束行
        :return:返回一个list，如[[u'qwer', u'data1', u'data2'], [2L, 40L, 30L]]
        """
        list1 = []
        if star_row == None and end_row == None:
            for x in self.sheet.rows:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_row != None and end_row == None:
            for x in list(self.sheet.rows)[star_row:]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_row == None and end_row != None:
            for x in list(self.sheet.rows)[:end_row]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_row != None and end_row != None:
            for x in list(self.sheet.rows)[star_row:end_row]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        return list1

    def openpyxl_Get_ColumnsText(self, star_column=None, end_column=None):
        """
        返回每列的内容，可以以切片进行处理
        1、star_column=None和end_column=None，返回sheet页全部内容
        2、star_column有值，end_column=None,返回的内容为columns[star_row:]
        3、star_column=None，end_column有值,返回的内容为columns[:end_row]
        4、star_column有值，end_column有值,返回的内容为columns[star_row:end_row]
        :param star_column: 起始列
        :param end_column: 结束列
        :return:返回一个list，如[[u'qwer', u'data1', u'data2'], [2L, 40L, 30L]]
        """
        list1 = []
        if star_column == None and end_column == None:
            for x in self.sheet.columns:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_column != None and end_column == None:
            for x in list(self.sheet.columns)[star_column:]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_column == None and end_column != None:
            for x in list(self.sheet.columns)[:end_column]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        elif star_column != None and end_column != None:
            for x in list(self.sheet.columns)[star_column:end_column]:
                list2 = []
                for y in x:
                    list2.append(y.value)
                list1.append(list2)
        return list1

    def openpyxl_Create_WorkBook(self, active=True, title=None):
        """
        创建一个工作表
        :param active: 默认新增一个sheet页面，active为False,不会新增一个默认sheet页
        :param title: 默认新增的sheet的名称为默认名称，title不等于None，则修改默认sheet的名称
        :return:
        """
        self.wb = Workbook()
        if active == True:
            self.sheet = self.wb.active
            #self.sheet = self.wb.get_active_sheet
            if title != None:
                self.sheet.title = title

    def openpyxl_Set_SheetName(self, NewName):
        """
        修改sheet的名称
        :param NewName:新的名称
        :return:
        """
        self.sheet.title = NewName

    def openpyxl_Set_CellCoordinate(self, CellCoordinate, Text):
        """
        单元格中写入内容
        :param CellCoordinate: 需要修改的单元格坐标，如"A1"
        :param Text: 修改的内容
        :return:
        """
        self.sheet[CellCoordinate] = Text

    def openpyxl_Set_CellNumber(self, Cell_Row, Cell_Column, Text):
        """
         单元格中写入内容
        :param Cell_Row: 目标单元格X轴坐标
        :param Cell_Column: 目标单元格Y轴坐标
        :param Text: 修改的内容
        :return:
        """
        self.sheet.cell(row=Cell_Row, column=Cell_Column).value = Text

    def openpyxl_Set_CellRows(self,Rows):
        """
        以行为单位进行插入
        :param Rows:数据格式为[["1", "2", "3"],["1", "2", "3"],["1", "2", "3"]]
        :return:
        """
        for Row in Rows:
            self.sheet.append(Row)

    def openpyxl_Set_CellFont_CellCoordinate(self, CellCoordinate, name=None, size=None, bold=None, italic=None, vertAlign=None, underline=None, strike=None, color=None):
        """
        设置单元格字体样式
        :param CellCoordinate:cell定位，如"A1"
        :param name:字体样式
        :param size:字体大小
        :param bold:是否为粗体
        :param italic:是否为斜体
        :param vertAlign:对齐方式 None、subscript（下标）、baseline（基线）、superscript（上标）
        :param underline:下划线，double、single、doubleAccounting、singleAccountin
        :param strike:删除线
        :param color:字体颜色
        :return:
        """
        self.sheet[CellCoordinate].font = Font(name=name, size=size, bold=bold, italic=italic, vertAlign=vertAlign, underline=underline, strike=strike, color=color)

    def openpyxl_Set_CellFont_CellNumber(self,Cell_Row,Cell_Column, name=None, size=None, bold=None, italic=None, vertAlign=None, underline=None, strike=None, color=None):
        """
        设置单元格字体样式
        :param Cell_Row: 单元格X轴坐标
        :param Cell_Column: 单元格y轴坐标
        :param name:字体样式
        :param size:字体大小
        :param bold:是否为粗体
        :param italic:是否为斜体
        :param vertAlign:对齐方式 None、subscript（下标）、baseline（基线）、superscript（上标）
        :param underline:下划线，double、single、doubleAccounting、singleAccountin
        :param strike:删除线
        :param color:字体颜色
        :return:
        """
        self.sheet.cell(row=Cell_Row, column=Cell_Column).font = Font(name=name, size=size, bold=bold, italic=italic, vertAlign=vertAlign, underline=underline, strike=strike, color=color)

    def openpyxl_Set_CellPatternFill_Cellcoordinate(self, CellCoordinate, fill_type=None, start_color= None, end_color=None):
        """
        单元格填充颜色
        :param CellCoordinate: 单元格坐标，如"A1"
        :param fill_type: 填充格式
        :param start_color: 填充开始颜色
        :param end_color: 填充结束颜色
        :return:
        """
        self.sheet[CellCoordinate].patterFill = PatternFill(fill_type=fill_type, start_color= start_color, end_color=end_color)

    def openpyxl_Set_CellPatternFill_CellNumber(self,Cell_Row,Cell_Column, fill_type=None, start_color= None, end_color=None):
        """
        单元格填充颜色
        :param Cell_Row: 单元格X轴坐标
        :param Cell_Column: 单元格y轴坐标
        :param fill_type: 填充格式
        :param start_color: 填充开始颜色
        :param end_color: 填充结束颜色
        :return:
        """
        self.sheet.cell(row=Cell_Row, column=Cell_Column).patterFill = PatternFill(fill_type=fill_type, start_color= start_color, end_color=end_color)

    def openpyxl_Set_CellAlignment_CellCoordinate(self, CellCorrdinate,horizontal=None, vertical=None, text_rotation=None, wrap_text=None, shrink_to_fit=None, indent=None):
        """
        单元格对齐方式
        :param CellCorrdinate: 单元格坐标，如"A1"
        :param horizontal:水平方向，right、center、centerContinuous、distributed、general、left、justify、fill
        :param vertical:上下方向，top、justify、distributed、center、bottom
        :param text_rotation:
        :param wrap_text:
        :param shrink_to_fit:
        :param indent:
        :return:
        """
        self.sheet[CellCorrdinate].alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text, shrink_to_fit=shrink_to_fit, indent=indent)

    def openpyxl_Set_CellAlignment_CellNumber(self, Cell_Row,Cell_Column, horizontal=None, vertical=None, text_rotation=None, wrap_text=None, shrink_to_fit=None, indent=None):
        """
        单元格对齐方式
        :param Cell_Row: 单元格X轴坐标
        :param Cell_Column: 单元格y轴坐标
        :param horizontal:水平方向，right、center、centerContinuous、distributed、general、left、justify、fill
        :param vertical:上下方向，top、justify、distributed、center、bottom
        :param text_rotation:
        :param wrap_text:
        :param shrink_to_fit:
        :param indent:
        :return:
        """
        self.sheet.cell(row=Cell_Row, column=Cell_Column).alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation, wrap_text=wrap_text, shrink_to_fit=shrink_to_fit, indent=indent)

    def openpyxl_Set_CellBroder_CellCoordinate(self, CellCorrdinate, border="left", border_style=None, color=None):
        """
        单元格边框样式
        :param CellCorrdinate: 单元格坐标，如"A1"
        :param border: 边框（leftright、top、bottom）
        :param border_style: 边框样式
        :param color: 边框样色
        :return:
        """
        if border == 'left':
            self.sheet[CellCorrdinate].border = Border(left=Side(border_style=border_style, color=color))
        elif border == 'right':
            self.sheet[CellCorrdinate].border = Border(right=Side(border_style=border_style, color=color))
        elif border == 'top':
            self.sheet[CellCorrdinate].border = Border(top=Side(border_style=border_style, color=color))
        elif border == 'bottom':
            self.sheet[CellCorrdinate].border = Border(bottom=Side(border_style=border_style, color=color))

    def openpyxl_Set_CellBroder_CellNumber(self, CellRow, CellColumn, border="left", border_style=None, color=None):
        """
        单元格表框样式
        :param Cell_Row: 单元格X轴坐标
        :param Cell_Column: 单元格y轴坐标
        :param CellCorrdinate: 单元格坐标，如"A1"
        :param border: 边框（leftright、top、bottom）
        :param border_style: 边框样式
        :param color: 边框样色
        :return:
        """
        if border == 'left':
            self.sheet.cell(row=CellRow, column=CellColumn).border = Border(left=Side(border_style=border_style, color=color))
        elif border == 'right':
            self.sheet.cell(row=CellRow, column=CellColumn).border = Border(right=Side(border_style=border_style, color=color))
        elif border == 'top':
            self.sheet.cell(row=CellRow, column=CellColumn).border = Border(top=Side(border_style=border_style, color=color))
        elif border == 'bottom':
            self.sheet.cell(row=CellRow, column=CellColumn).border = Border(bottom=Side(border_style=border_style, color=color))

    def openpyxl_Set_Rowheight(self, Row, height):
        """
        设置行高
        :param Row:行数
        :param height: 行高
        :return:
        """
        self.sheet.row_dimensions[Row].height = height
    def openpyxl_Set_ColumnWidth(self, Column, width):
        """
        设置列宽
        :param Column:列数
        :param width: 列宽
        :return:
        """
        self.sheet.column_dimensions[Column].width = width

    def openpyxl_Get_MergedCell(self):
        """
        获取sheet页中所有单元格
        :return:
        """
        return self.sheet.merged_cell_ranges
        #返回单元格范围['E2:E3', 'E4:E7', 'D8:E8']
    def openpyxl_Get_MergedCell_Coordinate(self):
        """
        获取单元格具体坐标
        :return:
        """
        return list(self.sheet.merged_cells)
        #返回单元格具体的坐标，如set(['E8', 'D8', 'E5', 'E4', 'E7', 'E6', 'E3', 'E2'])
    def openpyxl_Add_MergedCell(self, star_cell, end_cell):
        """
        合并单元格
        :param star_cell: 起始坐标位置
        :param end_cell: 结束坐标位置
        :return:
        """
        self.sheet.merged_cells(star_cell + ":" + end_cell)

    def openpyxl_Delete_MergedCell(self, star_cell, end_cell):
        """
        取消合并单元格
        :param star_cell: 起始坐标位置
        :param end_cell: 结束坐标位置
        :return:
        """
        self.sheet.unmerge_cells(star_cell + ":" + end_cell)
    def pp(self):
        self.sheet.merge_cells('B7:B8')
        self.sheet.unmerge_cells('B7:B8')
        print  self.sheet.merged_cell_ranges
        #print self.sheet.merged_cells

        #self.sheet.unmerge_cell_ranges('E2:E3')


